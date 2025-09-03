import json
import re
import sys
import shutil
from pathlib import Path
from datetime import datetime
import pandas as pd # pyright: ignore[reportMissingModuleSource]

from openpyxl.styles import Font, PatternFill, Alignment # pyright: ignore[reportMissingModuleSource]
from openpyxl.formatting.rule import FormulaRule # pyright: ignore[reportMissingModuleSource]
from openpyxl.drawing.image import Image as XLImage # pyright: ignore[reportMissingModuleSource]

ROOT = Path(__file__).resolve().parents[1]
MOCK = ROOT / "mock_data"
OUT  = ROOT / "output"
MAP  = ROOT / "framework_map.csv"
ASSETS = ROOT / "assets"                 # optional: place a logo at assets/logo.png
EVIDENCE_DIR = ROOT / "evidence"         # place screenshots here
EVIDENCE_MAP = EVIDENCE_DIR / "evidence_map.csv"  # optional metadata override

OUT.mkdir(parents=True, exist_ok=True)

# --------- Cover Sheet Config ---------
COVER_INFO = {
    "Project": "EDS – GCP Audit Evidence Demo",
    "Scope": "Demonstrate automated export of GCP security findings & SIEM events into audit-ready Excel for compliance mapping.",
    "Data Sources": "Mock SCC-like findings + Mock SIEM events (synthetic); mapping via framework_map.csv",
    "Owner / Contact": "Eagle Defense Systems LLC (EDS)",
    "Frameworks Included": [
        "NIST 800-53 Rev. 5",
        "CMMC (select practices)",
        "FedRAMP (alignment references)",
        "SOC 2 (TSC CC series)",
        "ISO/IEC 27001:2022 (Annex A)"
    ]
}
# --------------------------------------

def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [re.sub(r"\s+", " ", c).strip().lower() for c in df.columns]
    return df

def load_json(path: Path) -> pd.DataFrame:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    df = pd.json_normalize(data)
    return normalize_cols(df)

def load_map_csv(path: Path) -> pd.DataFrame:
    for enc in ("utf-8-sig", "utf-8", "utf-16", "cp1252"):
        try:
            df = pd.read_csv(path, dtype=str, encoding=enc, engine="python", on_bad_lines="skip")
            df = normalize_cols(df)
            df = df.map(lambda x: x.strip() if isinstance(x, str) else x)
            if "rule_id" in df.columns:
                df["rule_id"] = df["rule_id"].astype(str).str.strip()
                return df
        except Exception:
            continue
    try:
        head = pd.read_csv(path, nrows=1, header=None, engine="python")
        print(f"[DEBUG] First row read (no header): {head.iloc[0].tolist()}", file=sys.stderr)
    except Exception as e:
        print(f"[DEBUG] Could not read CSV for diagnostics: {e}", file=sys.stderr)
    raise SystemExit(f"[ERROR] 'rule_id' column not found in {path}. "
                     f"Please ensure the header includes 'rule_id' (e.g., rule_id,nist_ref,cmmc_ref,fedramp_ref,soc2_ref,iso27001_ref,excel_tab)")

def exec_summary(findings_df, siem_df, scuba_df):
    s1 = findings_df.groupby("severity", dropna=False).size().rename("count").reset_index()
    s1["source"] = "GCP Findings"
    s2 = siem_df.groupby("severity", dropna=False).size().rename("count").reset_index()
    s2["source"] = "SIEM"
    s3 = scuba_df.groupby("severity", dropna=False).size().rename("count").reset_index()
    s3["source"] = "ScubaGoggles"
    summ = pd.concat([s1, s2, s3], ignore_index=True)
    top_rules_f = findings_df.groupby("rule_id", dropna=False).size().sort_values(ascending=False).head(5).rename("count").reset_index()
    top_rules_s = siem_df.groupby("rule_id", dropna=False).size().sort_values(ascending=False).head(5).rename("count").reset_index()
    return summ, top_rules_f, top_rules_s

def strip_tz(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        # robust check for tz-aware dtypes
        try:
            if isinstance(df[col].dtype, pd.DatetimeTZDtype):
                df[col] = df[col].dt.tz_convert("UTC").dt.tz_localize(None)
        except Exception:
            pass
    return df

# ---------- Excel Formatting Helpers --------------------

def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
            except Exception:
                val = ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

def style_header(ws, header_row=1):
    bold = Font(bold=True)
    fill = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")
    for cell in ws[header_row]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = ws["A2"]
    ws.auto_filter.ref = ws.dimensions

def format_dates(ws, header_row=1):
    headers = [c.value for c in ws[header_row]]
    for idx, name in enumerate(headers, start=1):
        if not isinstance(name, str):
            continue
        if re.search(r"(time|date)", name.lower()):
            for r in ws.iter_rows(min_row=header_row+1, min_col=idx, max_col=idx):
                for cell in r:
                    try:
                        if hasattr(cell.value, "year"):
                            cell.number_format = "yyyy-mm-dd hh:mm:ss"
                    except Exception:
                        pass

def add_severity_heat(ws, header_row=1):
    headers = [c.value for c in ws[header_row]]
    try:
        sev_idx = [h.lower() for h in headers].index("severity") + 1
    except Exception:
        return
    last_row = ws.max_row
    start_cell = ws.cell(row=header_row+1, column=sev_idx).coordinate
    end_cell = ws.cell(row=last_row, column=sev_idx).coordinate
    rng = f"{start_cell}:{end_cell}"
    col_letter = re.match(r"([A-Z]+)", start_cell).group(1)
    rules = [
        ("UPPER({cell})=\"HIGH\"",   "FFFFC7CE"),
        ("UPPER({cell})=\"MEDIUM\"", "FFFFE4B5"),
        ("UPPER({cell})=\"LOW\"",    "FFC6EFCE"),
    ]
    for formula, color in rules:
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[formula.format(cell=f"{col_letter}{header_row+1}")],
                stopIfTrue=True,
                fill=PatternFill(start_color=color, end_color=color, fill_type="solid")
            )
        )

def beautify_sheet(ws, tab_color=None):
    style_header(ws)
    autosize_columns(ws)
    format_dates(ws)
    add_severity_heat(ws)
    if tab_color:
        ws.sheet_properties.tabColor = tab_color

def add_branding(ws, run_ts: str, top_right_anchor="H1"):
    try:
        ws.oddFooter.center.text = f"Eagle Defense Systems LLC – GCP Audit Evidence Demo | Generated (UTC): {run_ts}"
    except Exception:
        pass
    logo_path = ASSETS / "logo.png"
    if logo_path.exists():
        try:
            img = XLImage(str(logo_path))
            img.height = 60
            img.width = 180
            ws.add_image(img, top_right_anchor)
        except Exception:
            pass

# -------------------- Cover Sheet --------------------

def add_cover_sheet(wb, run_ts: str):
    ws = wb.create_sheet("Cover", 0)  # first tab
    ws.sheet_properties.tabColor = "9BBB59"  # greenish
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    title = "GCP Audit Evidence Package"
    ws["A1"].value = title
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:B1")

    # key/value rows
    row = 3
    kv_pairs = [
        ("Project", COVER_INFO.get("Project", "")),
        ("Scope", COVER_INFO.get("Scope", "")),
        ("Run Timestamp (UTC)", run_ts),
        ("Data Sources", COVER_INFO.get("Data Sources", "")),
        ("Owner / Contact", COVER_INFO.get("Owner / Contact", "")),
    ]
    for k, v in kv_pairs:
        ws[f"A{row}"].value = k
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].value = v
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        row += 2

    # Framework bullets
    ws[f"A{row}"].value = "Frameworks Included"
    ws[f"A{row}"].font = Font(bold=True)
    frameworks = COVER_INFO.get("Frameworks Included", [])
    r0 = row
    for i, fw in enumerate(frameworks, start=1):
        ws[f"B{row}"].value = f"• {fw}"
        row += 1
    if row == r0:
        ws[f"B{row}"].value = "• (none specified)"

    # branding/logo
    add_branding(ws, run_ts, top_right_anchor="B2")

    return ws

# ---------- Coverage Matrix helpers ----------

def _split_controls(cell: str) -> list[str]:
    if not isinstance(cell, str) or not cell.strip():
        return []
    parts = [p.strip() for p in re.split(r"[;,]", cell)]
    cleaned = []
    for p in parts:
        p = re.sub(r"^[A-Za-z/ ]+[\-–—]\s*", "", p).strip()
        if re.search(r"[A-Za-z]\.?[A-Za-z0-9\-\.]+", p):
            cleaned.append(p)
    return [c for c in cleaned if c]

def build_coverage_df(fmap: pd.DataFrame) -> pd.DataFrame:
    cols = [
        ("NIST 800-53", "nist_ref"),
        ("CMMC", "cmmc_ref"),
        ("FedRAMP", "fedramp_ref"),
        ("SOC 2", "soc2_ref"),
        ("ISO/IEC 27001", "iso27001_ref"),
    ]
    rows = []
    for fw_name, col in cols:
        if col not in fmap.columns:
            rows.append((fw_name, 0, 0))
            continue
        controls = set()
        rules_mapped = 0
        for _, v in fmap[col].fillna("").items():
            items = _split_controls(v)
            if items:
                rules_mapped += 1
                controls.update(items)
        rows.append((fw_name, len(controls), rules_mapped))
    cov = pd.DataFrame(rows, columns=["framework", "unique_controls", "rules_mapped"])
    cov = cov.sort_values("framework").reset_index(drop=True)
    return cov

# ---------- Evidence Log helpers ----------

_IMG_EXTS = {".png", ".jpg", ".jpeg", ".webp"}

def _infer_ac_from_name(name: str) -> str | None:
    """
    Expect filenames like:
      AC.L2-3.1.1_iam_roles.png
      AC.L2-3.1.22_bucket_permissions.jpg
    Returns 'AC.L2-3.1.1' etc. or None.
    """
    m = re.match(r"^(AC[\.\w\-]+)", name, flags=re.IGNORECASE)
    return m.group(1) if m else None

def build_evidence_df(evidence_root: Path) -> pd.DataFrame:
    rows = []
    if not evidence_root.exists():
        return pd.DataFrame(columns=["ac_requirement","screenshot_filename","description","control_evidence","relative_path"])

    for p in evidence_root.iterdir():
        if not p.is_file():
            continue
        if p.suffix.lower() not in _IMG_EXTS:
            continue
        ac = _infer_ac_from_name(p.stem)
        rows.append({
            "ac_requirement": ac or "",
            "screenshot_filename": p.name,
            "description": "",
            "control_evidence": "",
            "relative_path": f"evidence/{p.name}"
        })

    df = pd.DataFrame(rows, columns=["ac_requirement","screenshot_filename","description","control_evidence","relative_path"])

# Merge optional metadata (accepts simple or rich schemas)
    if EVIDENCE_MAP.exists():
        try:
            meta = pd.read_csv(EVIDENCE_MAP, dtype=str)
            meta = normalize_cols(meta)

            # Map common aliases to the internal schema
            # simple schema: file,control,description
            # rich schema  : screenshot_filename, ac_requirement, description, control_evidence
            rename = {
                "file": "screenshot_filename",
                "screenshot": "screenshot_filename",
                "control": "ac_requirement",
                "ac": "ac_requirement",
                "ac_id": "ac_requirement",
            }
            meta.rename(columns=rename, inplace=True)

            # Ensure required columns exist (fill missing AC as empty to allow filename-only rows)
            if "screenshot_filename" not in meta.columns:
                raise ValueError("evidence_map.csv must include a 'file' or 'screenshot_filename' column")
            if "ac_requirement" not in meta.columns:
                meta["ac_requirement"] = ""

            # Keep only supported columns
            keep = [c for c in ("ac_requirement", "screenshot_filename", "description", "control_evidence") if c in meta.columns]
            meta = meta[keep]

            # Merge—by AC + filename when AC provided; filename-only rows still align via empty AC
            df = df.merge(meta, on=["ac_requirement", "screenshot_filename"], how="left", suffixes=("", "_meta"))

            # Prefer metadata text if present
            for col in ("description", "control_evidence"):
                if f"{col}_meta" in df.columns:
                    df[col] = df[f"{col}_meta"].combine_first(df[col])
                    df.drop(columns=[f"{col}_meta"], inplace=True)

        except Exception as e:
            print(f"[WARN] Failed to merge evidence_map.csv: {e}", file=sys.stderr)


    # Sort nicely: AC then filename
    if "ac_requirement" in df.columns:
        df = df.sort_values(["ac_requirement","screenshot_filename"], na_position="last").reset_index(drop=True)
    return df
# ----------- Transform Scuba Data ------------

def transform_scuba_data(scuba_raw: pd.DataFrame) -> pd.DataFrame:
    """Transform ScubaGoggles asset data into findings format."""
    findings = []
    
    if 'resources' in scuba_raw.columns:
        resources = scuba_raw['resources'].iloc[0]  # Get the resources array
        
        for resource in resources:
            resource_name = resource.get('name', 'Unknown')
            resource_type = resource.get('type', 'Unknown')
            project_id = resource.get('project_id', 'Unknown')
            
            # Check for public access (security finding)
            if 'iam' in resource and 'bindings' in resource['iam']:
                for binding in resource['iam']['bindings']:
                    if 'allUsers' in binding.get('members', []):
                        findings.append({
                            'rule_id': 'SCUBA_PUBLIC_ACCESS',
                            'resource_name': resource_name,
                            'resource_type': resource_type,
                            'project_id': project_id,
                            'severity': 'HIGH',
                            'description': f'Public access detected on {resource_type}',
                            'role': binding.get('role'),
                            'finding_class': 'MISCONFIGURATION',
                            'event_time': datetime.utcnow().isoformat()
                        })
            
            # Check for admin roles (compliance finding)
            if 'bindings' in resource:
                for binding in resource['bindings']:
                    if 'Admin' in binding.get('role', ''):
                        findings.append({
                            'rule_id': 'SCUBA_ADMIN_ROLE',
                            'resource_name': resource_name,
                            'resource_type': resource_type,
                            'project_id': project_id,
                            'severity': 'MEDIUM',
                            'description': f'Administrative role assignment detected',
                            'role': binding.get('role'),
                            'finding_class': 'COMPLIANCE_CHECK',
                            'event_time': datetime.utcnow().isoformat()
                        })
    
    return pd.DataFrame(findings)


# -------------------- Main --------------------

def main():
    findings = load_json(MOCK / "findings.json")
    siem = load_json(MOCK / "siem_events.json")
    
    # Load ScubaGoggles data and transform it
    scuba_raw = load_json(MOCK / "scuba_export.json")
    scuba = transform_scuba_data(scuba_raw) # pyright: ignore[reportUndefinedVariable]
    
    fmap = load_map_csv(MAP)

    for name, df in [("findings.json", findings), ("siem_events.json", siem), ("scuba_export.json", scuba)]:
        if "rule_id" not in df.columns:
            raise SystemExit(f"[ERROR] 'rule_id' missing in {name}. Columns: {list(df.columns)}")

    for df in (findings, siem, scuba):
        df["rule_id"] = df["rule_id"].astype(str).str.strip()
        if "event_time" in df.columns:
            df["event_time"] = pd.to_datetime(df["event_time"], errors="coerce", utc=True)
        if "severity" not in df.columns:
            df["severity"] = "UNKNOWN"

    findings_m = strip_tz(findings.merge(fmap, on="rule_id", how="left"))
    siem_m = strip_tz(siem.merge(fmap, on="rule_id", how="left"))
    scuba_m = strip_tz(scuba.merge(fmap, on="rule_id", how="left"))

    summary_counts, top_rules_f, top_rules_s = exec_summary(findings_m, siem_m, scuba_m)

    # Timestamped run folder
    run_ts = datetime.utcnow().strftime("%Y-%m-%d_%H%M%S")
    run_dir = OUT / run_ts
    run_dir.mkdir(parents=True, exist_ok=True)
    out_path = run_dir / "audit_report.xlsx"

    # Build evidence listing & copy files into run bundle
    evidence_df = build_evidence_df(EVIDENCE_DIR)
    bundle_evid_dir = run_dir / "evidence"
    if not evidence_df.empty:
        bundle_evid_dir.mkdir(exist_ok=True)
        for _, r in evidence_df.iterrows():
            src = EVIDENCE_DIR / r["screenshot_filename"]
            dst = bundle_evid_dir / r["screenshot_filename"]
            try:
                if src.exists():
                    shutil.copy2(src, dst)
            except Exception as e:
                print(f"[WARN] Could not copy evidence {src} -> {dst}: {e}", file=sys.stderr)

    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        # 0) Cover
        add_cover_sheet(xl.book, run_ts)

        # 1) Exec Summary
        summary_counts.to_excel(xl, sheet_name="Exec_Summary", index=False, startrow=1)
        ws = xl.sheets["Exec_Summary"]
        ws.cell(row=1, column=1).value = "Counts by Severity (Findings + SIEM + ScubaGoggles)"
        startrow = len(summary_counts) + 4
        top_rules_f.to_excel(xl, sheet_name="Exec_Summary", index=False, startrow=startrow)
        ws.cell(row=startrow, column=1).value = "Top Rules (Findings)"
        startrow = startrow + len(top_rules_f) + 3
        top_rules_s.to_excel(xl, sheet_name="Exec_Summary", index=False, startrow=startrow)
        ws.cell(row=startrow, column=1).value = "Top Rules (SIEM)"
        beautify_sheet(ws, tab_color="FFC000")
        add_branding(ws, run_ts)

        # 2) Evidence Log (if any evidence exists)
        if not evidence_df.empty:
            # Create hyperlink column to the bundled evidence path
            df_to_write = evidence_df.copy()
            # OpenPyXL hyperlinks work when we actually assign after write; we'll write plain cells, then convert
            df_to_write.to_excel(xl, sheet_name="Evidence_Log", index=False)
            ws_ev = xl.sheets["Evidence_Log"]
            beautify_sheet(ws_ev, tab_color="92D050")  # green
            # add hyperlinks (relative to workbook location)
            headers = [c.value for c in ws_ev[1]]
            link_col = headers.index("relative_path") + 1 if "relative_path" in headers else None
            file_col = headers.index("screenshot_filename") + 1 if "screenshot_filename" in headers else None
            if link_col and file_col:
                for row_idx in range(2, ws_ev.max_row + 1):
                    rel = ws_ev.cell(row=row_idx, column=link_col).value
                    disp = ws_ev.cell(row=row_idx, column=file_col).value
                    if rel and disp:
                        cell = ws_ev.cell(row=row_idx, column=file_col)
                        cell.hyperlink = rel  # hyperlink to file in ./evidence/
                        cell.style = "Hyperlink"
                # hide helper path column but keep in sheet
                ws_ev.column_dimensions[ws_ev.cell(row=1, column=link_col).column_letter].hidden = True

        # 3) Detail sheets
        findings_m.to_excel(xl, sheet_name="Findings_Detail", index=False)
        siem_m.to_excel(xl, sheet_name="SIEM_Incidents", index=False)
        scuba_m.to_excel(xl, sheet_name="ScubaGoggles_Assessment", index=False)
        fmap.to_excel(xl, sheet_name="Framework_Mapping", index=False)

        beautify_sheet(xl.sheets["Findings_Detail"], tab_color="5B9BD5")
        beautify_sheet(xl.sheets["SIEM_Incidents"],  tab_color="ED7D31")
        beautify_sheet(xl.sheets["ScubaGoggles_Assessment"], tab_color="FF6B35")  # Orange
        beautify_sheet(xl.sheets["Framework_Mapping"], tab_color="70AD47")

        # Optional: add a simple Coverage summary tab
        try:
            cov = build_coverage_df(fmap)
            if not cov.empty:
                cov.to_excel(xl, sheet_name="Coverage_Summary", index=False)
                beautify_sheet(xl.sheets["Coverage_Summary"], tab_color="8FAADC")
        except Exception as e:
            print(f"[WARN] Coverage summary failed: {e}", file=sys.stderr)

    print(f"Wrote Excel report to: {out_path}")

    # Zip the run folder
    zip_path = shutil.make_archive(base_name=str(run_dir), format="zip", root_dir=str(run_dir))
    print(f"Packaged auditor bundle: {zip_path}")

if __name__ == "__main__":
    main()
